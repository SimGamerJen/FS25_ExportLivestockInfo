#!/usr/bin/env python3
# Export FS2025 livestock (genetics, pregnancy, summaries, Excel/CSV)
#
# - Reads ONLY <save>/placeables.xml (no game files modified)
# - Animals + Fetuses tables, plus Summary (by shed/species/breed)
# - Derived: age_days, age_years, preg_due_date
# - Country mapping: RealisticLivestock.lua AREA_CODES (or --country-map JSON)
# - Respects gameSettings.xml <modsDirectoryOverride>
# - Filters: --species cows,sheep,pigs  --farmid N
# - Batch mode: --all-saves (exports every savegame*)
# - Excel output (--xlsx) with proper numeric/date types; pre-1900 dates stay as text
#
# Usage (run from your FS25 folder):
#   py export_livestock_to_csv.py --save savegame1 --xlsx --verbose

import argparse
import csv
import json
import os
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, timedelta
from typing import Dict, List, Optional, Tuple

# ---------- Optional Excel support ----------
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

# ---------- Column definitions ----------
ANIMAL_COLUMNS = [
    "placeable_id",
    "current_shed",
    "shed_type",
    "species",
    "breed",
    "FarmID",
    "unique_id",
    "name",
    "sex",
    "age",           # months (as stored)
    "age_days",      # derived ~ months * 30
    "age_years",     # derived = months / 12
    "weight",
    "animal_health",
    "animal_gen_metabolism",
    "animal_gen_quality",
    "animal_gen_health",
    "animal_gen_fertility",
    "animal_gen_productivity",
    "birthday_day",
    "birthday_month",
    "birthday_year",
    "country",
    "country_name",
    "country_iso",
    "is_parent",
    "is_pregnant",
    "preg_day",
    "preg_month",
    "preg_year",
    "preg_duration",
    "preg_due_date",
    "preg_fetus_count",
    # placeholders (not in placeables.xml, kept for compatibility)
    "purchase_date",
    "purchase_price",
]

FETUS_COLUMNS = [
    "mother_unique_id",
    "current_shed",
    "shed_type",
    "species",
    "mother_breed",
    "fetus_index",
    "sex",
    "breed",
    "health",
    "gen_metabolism",
    "gen_quality",
    "gen_health",
    "gen_fertility",
    "gen_productivity",
    "due_date",
    "preg_day",
    "preg_month",
    "preg_year",
    "preg_duration",
    "FarmID",
    "country",
    "country_name",
    "country_iso",
]

SUMMARY_COLUMNS = [
    "current_shed",
    "shed_type",
    "species",
    "breed",
    "animals_count",
    "pregnant_count",
    "avg_age_months",
    "avg_age_years",
    "avg_health",
    "avg_gen_metabolism",
    "avg_gen_quality",
    "avg_gen_health",
    "avg_gen_fertility",
    "avg_gen_productivity",
    "total_fetuses",
]

# ---------- XLSX casting helpers (numbers & dates; pre-1900 stays text) ----------

ANIMAL_INT_COLS = {
    "birthday_day","birthday_month","birthday_year",
    "preg_day","preg_month","preg_year","preg_duration","preg_fetus_count","age_days",
    # NEW: cast IDs/codes to integers
    "FarmID", "unique_id", "country",
}

ANIMAL_FLOAT_COLS = {
    "age","age_years","weight","animal_health",
    "animal_gen_metabolism","animal_gen_quality","animal_gen_health",
    "animal_gen_fertility","animal_gen_productivity",
}
ANIMAL_DATE_COLS = {"preg_due_date"}

FETUS_INT_COLS = {
    "fetus_index","preg_day","preg_month","preg_year","preg_duration",
    # NEW: mirror numeric casting for fetus sheet
    "FarmID", "country", "mother_unique_id",
}

FETUS_FLOAT_COLS = {
    "health","gen_metabolism","gen_quality","gen_health","gen_fertility","gen_productivity"
}
FETUS_DATE_COLS = {"due_date"}

SUMMARY_INT_COLS = {"animals_count","pregnant_count","total_fetuses"}
SUMMARY_FLOAT_COLS = {
    "avg_age_months","avg_age_years","avg_health",
    "avg_gen_metabolism","avg_gen_quality","avg_gen_health",
    "avg_gen_fertility","avg_gen_productivity",
}
SUMMARY_DATE_COLS = set()

def _to_int_or_none(v: str):
    if v is None or v == "": return None
    try:
        f = float(v); i = int(round(f))
        return i
    except Exception:
        return None

def _to_float_or_none(v: str):
    if v is None or v == "": return None
    try:
        return float(v)
    except Exception:
        return None

def _excel_date_or_text(v: str):
    """Return a real date object for Excel if year >= 1900; else return the original text or None."""
    if v is None or v == "": return None
    try:
        d = date.fromisoformat(v)  # YYYY-MM-DD
        if d.year >= 1900:
            return d          # Excel-valid date -> real date object
        else:
            return v          # pre-1900 -> keep as TEXT to avoid negative serials / #####
    except Exception:
        return v              # not a valid ISO date -> keep as TEXT

def _cast_for_sheet(row: dict, columns: list, int_cols: set, float_cols: set, date_cols: set):
    out = []
    for col in columns:
        val = row.get(col, "")
        if col in date_cols:
            out.append(_excel_date_or_text(val))
        elif col in int_cols:
            out.append(_to_int_or_none(val))
        elif col in float_cols:
            out.append(_to_float_or_none(val))
        else:
            out.append(val if val != "" else None)
    return out

# ---------- Tiny XML helpers ----------
def _basename_or(val: str, fallback: str = "") -> str:
    try: return os.path.basename(val) if val else fallback
    except Exception: return fallback

def _text(elem: Optional[ET.Element]) -> str:
    return (elem.text or "").strip() if elem is not None and elem.text else ""

def _get_attr(elem: Optional[ET.Element], name: str, default: str = "") -> str:
    return elem.get(name, default) if elem is not None else default

def _child(elem: Optional[ET.Element], name: str) -> Optional[ET.Element]:
    return elem.find(name) if elem is not None else None

def _genetics(elem: Optional[ET.Element]) -> Dict[str, str]:
    g = _child(elem, "genetics")
    return {
        "metabolism": _get_attr(g, "metabolism"),
        "quality": _get_attr(g, "quality"),
        "health": _get_attr(g, "health"),
        "fertility": _get_attr(g, "fertility"),
        "productivity": _get_attr(g, "productivity"),
    }

# ---------- Species detection & normalization ----------
SPECIES_KEYS = {
    "cow": "cows", "cattle": "cows", "beef": "cows",
    "sheep": "sheep", "lamb": "sheep", "ovine": "sheep",
    "pig": "pigs", "hog": "pigs", "swine": "pigs", "porcine": "pigs",
    "chicken": "chickens", "hen": "chickens", "poultry": "chickens",
    "goat": "goats", "caprine": "goats",
    "horse": "horses", "equine": "horses",
}

def infer_species(placeable: ET.Element, shed_file: str, animal: ET.Element) -> str:
    t = _get_attr(animal, "type") or _get_attr(animal, "animalType")
    if t:
        tl = t.lower()
        for k, v in SPECIES_KEYS.items():
            if k in tl: return v
    pt = _text(_child(placeable, "type")).lower()
    for k, v in SPECIES_KEYS.items():
        if k and k in pt: return v
    lf = (shed_file or "").lower()
    for k, v in SPECIES_KEYS.items():
        if k and k in lf: return v
    return ""

def normalize_species_list(s: str) -> List[str]:
    out = []
    for raw in s.split(","):
        k = raw.strip().lower()
        if not k: continue
        out.append(k if k in SPECIES_KEYS.values() else SPECIES_KEYS.get(k, k))
    seen, res = set(), []
    for v in out:
        if v not in seen:
            seen.add(v); res.append(v)
    return res

# ---------- Country mapping (Realistic Livestock) ----------
def _str_to_bool(s: str) -> bool:
    return str(s).strip().lower() in ("1", "true", "yes", "on")

def find_mods_dir(fs_root: str, verbose: bool = False) -> str:
    settings_path = os.path.join(fs_root, "gameSettings.xml")
    default_dir = os.path.join(fs_root, "mods")
    if not os.path.isfile(settings_path):
        if verbose: print(f"[info] gameSettings.xml not found; using default mods dir: {default_dir}")
        return default_dir
    try:
        root = ET.parse(settings_path).getroot()
    except Exception as e:
        if verbose: print(f"[warn] could not parse gameSettings.xml: {e}")
        return default_dir
    node = root.find("modsDirectoryOverride")
    if node is None:
        for child in root:
            if child.tag.lower() == "modsdirectoryoverride":
                node = child; break
    if node is None:
        if verbose: print(f"[info] modsDirectoryOverride not present; using default mods dir: {default_dir}")
        return default_dir
    active = node.get("active") or ""
    directory = node.get("directory") or ""
    if not directory:
        dc = node.find("directory")
        if dc is not None and dc.text: directory = dc.text.strip()
    if _str_to_bool(active) and directory:
        directory = os.path.expanduser(os.path.expandvars(directory))
        if verbose: print(f"[info] modsDirectoryOverride active -> {directory}")
        return directory
    if verbose: print(f"[info] modsDirectoryOverride inactive; using default mods dir: {default_dir}")
    return default_dir

def _looks_like_rl_mod(name: str) -> bool:
    n = name.lower()
    return ("realistic" in n and "livestock" in n) or n.startswith("fs25_realisticlivestock")

def _brace_body(text: str, header: str) -> Optional[str]:
    i = text.find(header)
    if i < 0: return None
    j = text.find("{", i)
    if j < 0: return None
    depth = 0
    for k in range(j, len(text)):
        c = text[k]
        if c == "{": depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                return text[j+1:k]
    return None

def _parse_area_codes_from_lua(lua_text: str) -> Tuple[Dict[str,str], Dict[str,str]]:
    body = _brace_body(lua_text, "AREA_CODES")
    if not body: return {}, {}
    name_map, iso_map = {}, {}
    for entry in re.finditer(r'\[\s*(\d+)\s*\]\s*=\s*{(.*?)}', body, flags=re.S):
        idx = entry.group(1); sub = entry.group(2)
        m_code = re.search(r'\["code"\]\s*=\s*"([^"]*)"', sub)
        m_name = re.search(r'\["country"\]\s*=\s*"([^"]*)"', sub)
        if m_name: name_map[idx] = m_name.group(1)
        if m_code: iso_map[idx]  = m_code.group(1)
    return name_map, iso_map

def _load_area_codes_from_rl_path(path: str, verbose: bool=False) -> Tuple[Dict[str,str], Dict[str,str]]:
    if os.path.isdir(path):
        for root, _dirs, files in os.walk(path):
            for fn in files:
                if fn.lower() == "realisticlivestock.lua":
                    lua_path = os.path.join(root, fn)
                    if verbose: print(f"[info] reading RL lua: {lua_path}")
                    try:
                        with open(lua_path, "r", encoding="utf-8", errors="ignore") as fh:
                            txt = fh.read()
                        return _parse_area_codes_from_lua(txt)
                    except Exception as e:
                        if verbose: print(f"[warn] failed to read RL lua: {e}")
    elif os.path.isfile(path) and path.lower().endswith(".zip"):
        try:
            with zipfile.ZipFile(path, "r") as zf:
                cands = [n for n in zf.namelist() if n.lower().endswith("/realisticlivestock.lua") or n.lower()=="realisticlivestock.lua"]
                for n in cands:
                    if verbose: print(f"[info] reading RL lua from zip: {path} -> {n}")
                    txt = zf.read(n).decode("utf-8", errors="ignore")
                    return _parse_area_codes_from_lua(txt)
        except Exception as e:
            if verbose: print(f"[warn] failed to inspect zip {path}: {e}")
    return {}, {}

def load_area_codes_from_rl(mods_dir: str, verbose: bool=False) -> Tuple[Dict[str,str], Dict[str,str], List[str]]:
    checked = []
    if not os.path.isdir(mods_dir):
        return {}, {}, checked
    entries = os.listdir(mods_dir)
    rl_candidates = [e for e in entries if _looks_like_rl_mod(e)]
    scan_order = rl_candidates + [e for e in entries if e not in rl_candidates]
    for entry in scan_order:
        p = os.path.join(mods_dir, entry)
        checked.append(p)
        names, isos = _load_area_codes_from_rl_path(p, verbose=verbose)
        if names:
            return names, isos, checked
    for entry in entries:
        p = os.path.join(mods_dir, entry)
        if os.path.isfile(p) and p.lower().endswith(".zip"):
            checked.append(p)
            try:
                with zipfile.ZipFile(p, "r") as zf:
                    lua_members = [n for n in zf.namelist() if n.lower().endswith("/realisticlivestock.lua") or n.lower()=="realisticlivestock.lua"]
                    if lua_members:
                        txt = zf.read(lua_members[0]).decode("utf-8", errors="ignore")
                        names, isos = _parse_area_codes_from_lua(txt)
                        if names:
                            return names, isos, checked
            except Exception:
                pass
    return {}, {}, checked

def _load_country_map_json(path: Optional[str]) -> Dict[str, str]:
    if not path: return {}
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        return {str(k): str(v) for k, v in data.items()}
    except Exception as e:
        print(f"[warn] failed to read country map '{path}': {e}", file=sys.stderr)
        return {}

def _country_lookup(code_raw: str, name_map: Dict[str,str], iso_map: Dict[str,str], json_map: Dict[str,str]) -> Tuple[str,str]:
    if not code_raw: return "", ""
    if code_raw in json_map: return json_map[code_raw], ""
    try:
        as_int = str(int(float(code_raw)))
        if as_int in json_map: return json_map[as_int], ""
    except Exception:
        pass
    try: key = str(int(float(code_raw)))
    except Exception: key = code_raw
    name = name_map.get(key, "")
    iso  = iso_map.get(key, "")
    if not name and code_raw in name_map:
        name = name_map[code_raw]; iso = iso_map.get(code_raw, iso)
    return (name or f"Unknown ({code_raw})", iso)

# ---------- Core parsing ----------
def safe_float(x: str) -> Optional[float]:
    try:
        return float(x)
    except Exception:
        return None

def fmt_num(x: Optional[float], nd=2) -> str:
    return f"{x:.{nd}f}" if x is not None else ""

def derive_age_days(months_str: str) -> str:
    m = safe_float(months_str)
    return str(int(round(m*30))) if m is not None else ""

def derive_age_years(months_str: str) -> str:
    m = safe_float(months_str)
    return f"{m/12.0:.2f}" if m is not None else ""

def derive_due_date(y: str, m: str, d: str, duration_str: str) -> str:
    try:
        if not (y and m and d and duration_str): return ""
        y2, m2, d2 = int(y), int(m), int(d)
        dur = int(float(duration_str))
        dt = date(y2, m2, d2) + timedelta(days=dur)
        return dt.isoformat()
    except Exception:
        return ""

def parse_placeables(placeables_path: str,
                     country_name_map: Dict[str,str],
                     country_iso_map: Dict[str,str],
                     json_override: Dict[str,str],
                     species_filter: Optional[List[str]],
                     farmid_filter: Optional[str],
                     verbose: bool=False) -> Tuple[List[Dict[str,str]], List[Dict[str,str]]]:
    try:
        tree = ET.parse(placeables_path)
    except Exception as e:
        print(f"[error] Failed to parse {placeables_path}: {e}", file=sys.stderr)
        return [], []
    root = tree.getroot()
    animals_rows: List[Dict[str,str]] = []
    fetuses_rows: List[Dict[str,str]] = []

    for plc in root.findall("placeable"):
        shed_file = plc.get("filename", "")
        placeable_id = plc.get("id", "") or plc.get("uniqueId", "")
        current_shed = _basename_or(shed_file, placeable_id or "Husbandry")
        shed_type = _text(_child(plc, "type")) or _basename_or(shed_file)

        ha = plc.find("husbandryAnimals")
        if ha is None: continue
        clusters = ha.find("clusters")
        if clusters is None: continue

        for animal in clusters.findall("animal"):
            species = infer_species(plc, shed_file, animal)
            if species_filter and species and (species not in species_filter):
                continue
            if farmid_filter and _get_attr(animal, "farmId") != str(farmid_filter):
                continue

            row = {k: "" for k in ANIMAL_COLUMNS}
            row["placeable_id"] = placeable_id
            row["current_shed"] = current_shed
            row["shed_type"]    = shed_type
            row["species"]      = species
            row["breed"]        = _get_attr(animal, "subType")
            row["FarmID"]       = _get_attr(animal, "farmId")
            row["unique_id"]    = _get_attr(animal, "id")
            row["name"]         = _get_attr(animal, "name")
            row["sex"]          = _get_attr(animal, "gender")
            row["age"]          = _get_attr(animal, "age")
            row["age_days"]     = derive_age_days(row["age"])
            row["age_years"]    = derive_age_years(row["age"])
            row["weight"]       = _get_attr(animal, "weight")
            row["is_parent"]    = _get_attr(animal, "isParent")
            row["is_pregnant"]  = _get_attr(animal, "isPregnant")

            # health + genetics
            row["animal_health"] = _get_attr(animal, "health")
            agen = _genetics(animal)
            row["animal_gen_metabolism"]   = agen["metabolism"]
            row["animal_gen_quality"]      = agen["quality"]
            row["animal_gen_health"]       = agen["health"]
            row["animal_gen_fertility"]    = agen["fertility"]
            row["animal_gen_productivity"] = agen["productivity"]

            # birthday & country
            bday = _child(animal, "birthday")
            if bday is not None:
                row["birthday_day"]   = _get_attr(bday, "day")
                row["birthday_month"] = _get_attr(bday, "month")
                row["birthday_year"]  = _get_attr(bday, "year")
                row["country"]        = _get_attr(bday, "country")
                cname, ciso           = _country_lookup(row["country"], country_name_map, country_iso_map, json_override)
                row["country_name"]   = cname
                row["country_iso"]    = ciso

            # pregnancy block (also build fetuses table)
            preg = _child(animal, "pregnancy")
            if preg is not None:
                row["preg_day"]      = _get_attr(preg, "day")
                row["preg_month"]    = _get_attr(preg, "month")
                row["preg_year"]     = _get_attr(preg, "year")
                row["preg_duration"] = _get_attr(preg, "duration")
                row["preg_due_date"] = derive_due_date(row["preg_year"], row["preg_month"], row["preg_day"], row["preg_duration"])

                pregnancies = _child(preg, "pregnancies")
                fetus_count = 0
                if pregnancies is not None:
                    idx = 0
                    for fetus in pregnancies.findall("pregnancy"):
                        idx += 1; fetus_count += 1
                        fg = _genetics(fetus)
                        frow = {k: "" for k in FETUS_COLUMNS}
                        frow["mother_unique_id"] = row["unique_id"]
                        frow["current_shed"]     = current_shed
                        frow["shed_type"]        = shed_type
                        frow["species"]          = species
                        frow["mother_breed"]     = row["breed"]
                        frow["fetus_index"]      = str(idx)
                        frow["sex"]              = _get_attr(fetus, "gender")
                        frow["breed"]            = _get_attr(fetus, "subType")
                        frow["health"]           = _get_attr(fetus, "health")
                        frow["gen_metabolism"]   = fg["metabolism"]
                        frow["gen_quality"]      = fg["quality"]
                        frow["gen_health"]       = fg["health"]
                        frow["gen_fertility"]    = fg["fertility"]
                        frow["gen_productivity"] = fg["productivity"]
                        frow["due_date"]         = row["preg_due_date"]
                        frow["preg_day"]         = row["preg_day"]
                        frow["preg_month"]       = row["preg_month"]
                        frow["preg_year"]        = row["preg_year"]
                        frow["preg_duration"]    = row["preg_duration"]
                        frow["FarmID"]           = row["FarmID"]
                        frow["country"]          = row["country"]
                        frow["country_name"]     = row["country_name"]
                        frow["country_iso"]      = row["country_iso"]
                        fetuses_rows.append(frow)
                row["preg_fetus_count"] = str(fetus_count)

            animals_rows.append(row)

    if verbose:
        print(f"[info] parsed: animals={len(animals_rows)}, fetuses={len(fetuses_rows)}")

    return animals_rows, fetuses_rows

# ---------- Summaries ----------
def _to_float(x: str) -> Optional[float]:
    try:
        if x == "" or x is None: return None
        return float(x)
    except Exception:
        return None

def summarize(animals_rows: List[Dict[str,str]]) -> List[Dict[str,str]]:
    # group by (shed, shed_type, species, breed)
    groups: Dict[Tuple[str,str,str,str], List[Dict[str,str]]] = {}
    for r in animals_rows:
        key = (r["current_shed"], r["shed_type"], r["species"], r["breed"])
        groups.setdefault(key, []).append(r)

    summary_rows: List[Dict[str,str]] = []
    for (shed, shed_type, species, breed), items in groups.items():
        count = len(items)
        pregnant = sum(1 for it in items if (it.get("is_pregnant") or "").lower() in ("true","1","yes","on"))
        # averages
        ages = [_to_float(it.get("age")) for it in items if _to_float(it.get("age")) is not None]
        healths = [_to_float(it.get("animal_health")) for it in items if _to_float(it.get("animal_health")) is not None]
        gm = [_to_float(it.get("animal_gen_metabolism")) for it in items if _to_float(it.get("animal_gen_metabolism")) is not None]
        gq = [_to_float(it.get("animal_gen_quality")) for it in items if _to_float(it.get("animal_gen_quality")) is not None]
        gh = [_to_float(it.get("animal_gen_health")) for it in items if _to_float(it.get("animal_gen_health")) is not None]
        gf = [_to_float(it.get("animal_gen_fertility")) for it in items if _to_float(it.get("animal_gen_fertility")) is not None]
        gp = [_to_float(it.get("animal_gen_productivity")) for it in items if _to_float(it.get("animal_gen_productivity")) is not None]
        total_fetuses = sum(int(it.get("preg_fetus_count") or "0") for it in items)

        def avg(arr):
            return sum(arr)/len(arr) if arr else None

        row = {k: "" for k in SUMMARY_COLUMNS}
        row["current_shed"] = shed
        row["shed_type"]    = shed_type
        row["species"]      = species
        row["breed"]        = breed
        row["animals_count"] = str(count)
        row["pregnant_count"] = str(pregnant)
        row["avg_age_months"] = fmt_num(avg(ages))
        row["avg_age_years"]  = fmt_num((avg(ages)/12.0) if ages else None)
        row["avg_health"]     = fmt_num(avg(healths))
        row["avg_gen_metabolism"]   = fmt_num(avg(gm))
        row["avg_gen_quality"]      = fmt_num(avg(gq))
        row["avg_gen_health"]       = fmt_num(avg(gh))
        row["avg_gen_fertility"]    = fmt_num(avg(gf))
        row["avg_gen_productivity"] = fmt_num(avg(gp))
        row["total_fetuses"]        = str(total_fetuses)

        summary_rows.append(row)

    summary_rows.sort(key=lambda r: (r["current_shed"], r["species"], r["breed"]))
    return summary_rows

# ---------- IO helpers ----------
def ensure_dir_for(path: str):
    d = os.path.dirname(path)
    if d and not os.path.isdir(d):
        os.makedirs(d, exist_ok=True)

def write_csv(path: str, cols: List[str], rows: List[Dict[str,str]]):
    ensure_dir_for(path)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in cols})

def _format_date_columns(ws, columns: List[str], date_cols: set):
    """Apply Excel date number_format to cells that actually contain dates."""
    if not date_cols: return
    for col_name in date_cols:
        if col_name not in columns: continue
        col_idx = columns.index(col_name) + 1  # 1-based
        for row in range(2, ws.max_row + 1):  # skip header
            cell = ws.cell(row=row, column=col_idx)
            # only format true date objects (text stays as text)
            if isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"

def write_xlsx(path: str, animals: List[Dict[str,str]], fetuses: List[Dict[str,str]], summary: List[Dict[str,str]]):
    if not OPENPYXL_OK:
        print("[warn] openpyxl not installed; writing CSVs instead.", file=sys.stderr)
        base = os.path.splitext(path)[0]
        write_csv(base + "_animals.csv", ANIMAL_COLUMNS, animals)
        write_csv(base + "_fetuses.csv", FETUS_COLUMNS, fetuses)
        write_csv(base + "_summary.csv", SUMMARY_COLUMNS, summary)
        return

    ensure_dir_for(path)
    wb = Workbook()

    # Animals
    ws = wb.active; ws.title = "Animals"
    ws.append(ANIMAL_COLUMNS)
    for r in animals:
        ws.append(_cast_for_sheet(r, ANIMAL_COLUMNS, ANIMAL_INT_COLS, ANIMAL_FLOAT_COLS, ANIMAL_DATE_COLS))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ANIMAL_COLUMNS))}{ws.max_row}"
    ws.freeze_panes = "A2"
    _format_date_columns(ws, ANIMAL_COLUMNS, ANIMAL_DATE_COLS)

    # Fetuses
    ws2 = wb.create_sheet("Fetuses")
    ws2.append(FETUS_COLUMNS)
    for r in fetuses:
        ws2.append(_cast_for_sheet(r, FETUS_COLUMNS, FETUS_INT_COLS, FETUS_FLOAT_COLS, FETUS_DATE_COLS))
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(FETUS_COLUMNS))}{ws2.max_row}"
    ws2.freeze_panes = "A2"
    _format_date_columns(ws2, FETUS_COLUMNS, FETUS_DATE_COLS)

    # Summary
    ws3 = wb.create_sheet("Summary")
    ws3.append(SUMMARY_COLUMNS)
    for r in summary:
        ws3.append(_cast_for_sheet(r, SUMMARY_COLUMNS, SUMMARY_INT_COLS, SUMMARY_FLOAT_COLS, SUMMARY_DATE_COLS))
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_COLUMNS))}{ws3.max_row}"
    ws3.freeze_panes = "A2"

    wb.save(path)

# ---------- Save iteration ----------
def list_saves(fs_root: str) -> List[str]:
    out = []
    for name in sorted(os.listdir(fs_root)):
        if not name.lower().startswith("savegame"): continue
        p = os.path.join(fs_root, name)
        if os.path.isdir(p) and os.path.isfile(os.path.join(p, "placeables.xml")):
            out.append(p)
    return out

# ---------- CLI & main ----------
def resolve_save_dir(save_arg: str) -> str:
    cand = os.path.abspath(save_arg)
    if os.path.isdir(cand): return cand
    return os.path.join(os.getcwd(), save_arg)

def run_for_save(save_dir: str,
                 out_csv: Optional[str],
                 xlsx_path: Optional[str],
                 summary_out_path: Optional[str],
                 country_json_path: Optional[str],
                 rl_path: Optional[str],
                 species_filter_list: Optional[List[str]],
                 farmid_filter: Optional[str],
                 verbose: bool):
    placeables_path = os.path.join(save_dir, "placeables.xml")
    if not os.path.isfile(placeables_path):
        print(f"[error] placeables.xml not found at: {placeables_path}", file=sys.stderr)
        return

    # Country maps
    json_override = _load_country_map_json(country_json_path)
    name_map: Dict[str,str] = {}
    iso_map: Dict[str,str]  = {}
    if not json_override:
        if rl_path:
            if verbose: print(f"[info] using RL path: {rl_path}")
            name_map, iso_map = _load_area_codes_from_rl_path(rl_path, verbose=verbose)
        else:
            mods_dir = find_mods_dir(os.getcwd(), verbose=verbose)
            if verbose: print(f"[info] scanning mods dir: {mods_dir}")
            load_names, load_isos, checked = load_area_codes_from_rl(mods_dir, verbose=verbose)
            name_map, iso_map = load_names, load_isos
            if verbose and not name_map:
                print("[warn] RealisticLivestock.lua with AREA_CODES not found.")
                if checked:
                    print("[info] paths checked:")
                    for p in checked:
                        print("  -", p)

    animals, fetuses = parse_placeables(
        placeables_path,
        name_map, iso_map, json_override,
        species_filter_list, farmid_filter, verbose=verbose
    )
    summary = summarize(animals)

    save_name = os.path.basename(save_dir.rstrip("/\\"))
    if xlsx_path:
        xlsx = xlsx_path
        if os.path.isdir(xlsx) or xlsx.endswith(os.sep):
            xlsx = os.path.join(xlsx, f"{save_name}_livestock.xlsx")
        write_xlsx(xlsx, animals, fetuses, summary)
        print(f"[ok] {save_name}: wrote {xlsx}")
    else:
        animals_out = out_csv or os.path.join(save_dir, "livestock.csv")
        fetuses_out = os.path.splitext(animals_out)[0] + "_fetuses.csv"
        summary_csv = summary_out_path or os.path.join(save_dir, "livestock_summary.csv")
        write_csv(animals_out, ANIMAL_COLUMNS, animals)
        write_csv(fetuses_out, FETUS_COLUMNS, fetuses)
        write_csv(summary_csv, SUMMARY_COLUMNS, summary)
        print(f"[ok] {save_name}: wrote {animals_out}, {fetuses_out}, {summary_csv}")

def main():
    ap = argparse.ArgumentParser(description="Export FS2025 livestock (+genetics, pregnancy, summaries, Excel).")
    g = ap.add_mutually_exclusive_group(required=False)
    g.add_argument("-s", "--save", default="savegame1", help="Savegame folder name or path (default: savegame1).")
    g.add_argument("--all-saves", action="store_true", help="Export every savegame* under the current folder.")

    ap.add_argument("-o", "--out", default=None,
                    help="Animals CSV path (default: <save>/livestock.csv). For --all-saves, ignored unless a directory.")
    ap.add_argument("--summary-out", default=None,
                    help="Summary CSV path (default: <save>/livestock_summary.csv).")
    ap.add_argument("--xlsx", nargs="?", const="", default=None,
                    help="Write Excel workbook (Animals, Fetuses, Summary). "
                         "If value is a directory, files are created inside it; "
                         "if omitted value (just --xlsx), defaults to <save>/<save>_livestock.xlsx")
    ap.add_argument("--country-map", default=None,
                    help="Optional JSON mapping for country codes to names (overrides RL).")
    ap.add_argument("--rl", default=None,
                    help="Path to the RealisticLivestock mod (folder or .zip). If omitted, auto-detect in mods dir.")
    ap.add_argument("--species", default=None,
                    help="Comma list to include (e.g., 'cows,sheep,pigs').")
    ap.add_argument("--farmid", default=None,
                    help="Only include animals with this FarmID (e.g., '2').")
    ap.add_argument("--verbose", action="store_true", help="Print discovery info.")

    args = ap.parse_args()
    fs_root = os.getcwd()

    species_filter_list = normalize_species_list(args.species) if args.species else None

    if args.all_saves:
        saves = list_saves(fs_root)
        if not saves:
            print("[info] no savegame* folders with placeables.xml found here.")
            return
        # If --xlsx has "", build per-save defaults; if it's a non-empty path, treat as directory unless it ends with .xlsx
        xlsx_is_dir = None
        xlsx_root = None
        if args.xlsx is not None:
            if args.xlsx == "":
                xlsx_is_dir = False
            else:
                if args.xlsx.endswith(os.sep) or (os.path.isdir(args.xlsx) or not args.xlsx.lower().endswith(".xlsx")):
                    xlsx_is_dir = True
                    xlsx_root = args.xlsx
        for sdir in saves:
            if args.xlsx is None:
                xlsx_path = None
            else:
                if xlsx_is_dir is True:
                    xlsx_path = xlsx_root
                elif xlsx_is_dir is False:
                    save_name = os.path.basename(sdir.rstrip("/\\"))
                    xlsx_path = os.path.join(sdir, f"{save_name}_livestock.xlsx")
                else:
                    xlsx_path = None
            run_for_save(
                sdir,
                args.out, xlsx_path, args.summary_out,
                args.country_map, args.rl,
                species_filter_list, args.farmid, args.verbose
            )
    else:
        save_dir = resolve_save_dir(args.save)
        if args.xlsx is None:
            xlsx_path = None
        elif args.xlsx == "":
            save_name = os.path.basename(save_dir.rstrip("/\\"))
            xlsx_path = os.path.join(save_dir, f"{save_name}_livestock.xlsx")
        else:
            xlsx_path = args.xlsx
        run_for_save(
            save_dir,
            args.out, xlsx_path, args.summary_out,
            args.country_map, args.rl,
            species_filter_list, args.farmid, args.verbose
        )

if __name__ == "__main__":
    main()
